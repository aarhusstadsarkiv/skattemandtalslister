# Small script for fixing mistakes in the excel sheets based on the findings

from fileinput import filename
import openpyxl

from pathlib import Path
from main import clean_number, clean_road_name

directory: Path = Path("Kios_gadenavne")
for vejviser in directory.glob("*.xlsx"):
    wb_temp = openpyxl.load_workbook(filename=vejviser)
    sheet = wb_temp.active
    i = 1  # used as an index indicator for rows. bit hacky, works perfectly
    for row in sheet.iter_rows(values_only=True):
        road_name: str = clean_road_name(str(row[1]))
        even_end: int = clean_number(row[5])
        if even_end == 99 and road_name == "aagade":
            cell = sheet.cell(row=i, column=6)
            cell.value = 16
        i += 1
    wb_temp.save(filename=vejviser)
